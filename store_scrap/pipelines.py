import os
from openpyxl import Workbook, load_workbook
from store_scrap.items import Product


class XlsxPipeline:

    wb = Workbook()

    def open_spider(self, spider):
        self.sheet_name = spider.name.title().replace('_', ' ')
        if self.sheet_name in self.wb.sheetnames:
            del self.wb[self.sheet_name]
        spider.ws = self.wb.create_sheet(title=self.sheet_name, index=0)
        spider.ws.append(list(Product.fields.keys()))

    def process_item(self, item, spider):
        spider.ws.append([item.get(field_name) for field_name in Product.fields.keys()])
        return item

    def close_spider(self, spider):
        excel_file_path = spider.settings['EXCEL_FILE_PATH']
        if not excel_file_path:
            return
        if os.path.exists(excel_file_path):
            prev_wb = load_workbook(excel_file_path)
            for sheet_name in prev_wb.sheetnames:
                if sheet_name not in self.wb.sheetnames:
                    self.wb.create_sheet(title=sheet_name, index=0)
                    for row in prev_wb[sheet_name].iter_rows(values_only=True):
                        self.wb[sheet_name].append(row)
            prev_wb.close()
        self.wb.save(spider.settings['EXCEL_FILE_PATH'])
